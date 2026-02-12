import os

import streamlit as st
import openpyxl
from icalendar import Calendar, Event
import datetime
import parsedatetime as pdt
import re

def format_selection(university):
    uni_format = None
    if university == "University of Newcastle Y1 SEM 2 2025":
        uni_format = {"preset": "University of Newcastle Y1 SEM 2 2025",
                      "uni": "UON",
                      "day": 1,
                      "date": 2,
                      "time": 3,
                      "campus": 5,
                      "group": 6,
                      "venue": 7,
                      "attendance": 8,
                      "type": 9,
                      "domain": 10,
                      "session": 11,
                      "staff": 12,
                      "updates": 13,
                      "fields": ["pbl"]
                    }

    if university == "University of Newcastle Y2 2026":
        uni_format = {"preset": "University of Newcastle Y2 2026",
                      "uni": "UON",
                      "day": 1,
                      "date": 2,
                      "time": 3,
                      "campus": 5,
                      "group": 6,
                      "venue": 7,
                      "attendance": 8,
                      "type": 9,
                      "domain": 10,
                      "session": 11,
                      "staff": 12,
                      "updates": 13,
                      "fields": ["pbl"]
                    }


    if university == "University of New England Y2 2026":
        uni_format = {
                      "preset": "University of New England Y2 2026",
                      "uni": "UNE",
                      "day": 1,
                      "date": 2,
                      "time": 3,
                      "campus": 6,
                      "group": 5,
                      "venue": 6,
                      "attendance": 10,
                      "type": 9,
                      "domain": 8,
                      "session": 7,
                      "staff": 11,
                      "updates": 12,
                      "fields": ["pbl", "clin", "comm"]
                      }
    return uni_format

def main_ui():
    st.set_page_config(page_title="JMPCalendarConverter")

    st.title("JMPCalendarConverter")
    st.text("Works well enough for now, I'm still ironing out bugs. Lmk if you notice anything @photogrudesh.")

    st.divider()

    if not st.toggle("Use your own file"):
        university = st.selectbox("Preset calendar - Year 2 only rn", ["University of Newcastle Y2 2026", "University of New England Y2 2026"])
    else:
        university = "Use your own file"

    file = None

    files = os.listdir()
    uoncals = []
    unecals = []

    for f in files:
        if f.endswith(".xlsx"):
            if f.startswith("CALCCCS"):
                uoncals.append(f)
            elif f.startswith("UNEARM"):
                unecals.append(f)

    uni_format = format_selection(university)
    if uni_format:
        if uni_format["uni"] == "UON":
            file = st.selectbox("Select file", uoncals)
        elif uni_format["uni"] == "UNE":
            file = st.selectbox("Select file", unecals)

    valid_custom = True

    if university == "Use your own file":
        file = st.file_uploader("Upload your own calendar file", type=None, accept_multiple_files=False)
        uni_format_select = st.radio("Format",
                              ["University of Newcastle Y2 2026", "University of New England Y2 2026", "custom"], horizontal=True)
        if uni_format_select == "custom":
            st.text("Be careful with these settings. If they're not perfect, your calendar will not convert. Enter the column that correlates with each attribute. Starting at 0 for A, left to right. Then set your title row. This is the row which contains the headings for each column. This number lines up exactly with Excel.")
            uni_format = {
             "preset": None,
             "uni": None,
             "day": None,
             "date": None,
             "time": None,
             "campus": None,
             "campus_code": None,
             "group": None,
             "venue": None,
             "attendance": None,
             "type": None,
             "domain": None,
             "session": None,
             "staff": None,
             "updates": None,
             "fields": None
             }

            col_cust1, col_cust2, col_cust3, col_cust4, col_cust5, col_cust6 = st.columns(6)

            try:
                with col_cust1:
                    uni_format["day"] =  st.selectbox("day", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["date"] = st.selectbox("date", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                with col_cust2:
                    uni_format["time"] =  st.selectbox("time", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["campus"] = st.selectbox("campus", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                with col_cust3:
                    uni_format["group"] =  st.selectbox("group", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["venue"] = st.selectbox("venue", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                with col_cust4:
                    uni_format["attendance"] =  st.selectbox("attendance", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["type"] = st.selectbox("type", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                with col_cust5:
                    uni_format["domain"] =  st.selectbox("domain", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["session"] = st.selectbox("session", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                with col_cust6:
                    uni_format["staff"] =  st.selectbox("staff", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
                    uni_format["updates"] = st.selectbox("updates", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20])
            except ValueError:
                st.error("Please enter a number")

            uni_format['fields'] = st.segmented_control("Pick the groups that apply to your semester", ["pbl", "clin", "comm"], selection_mode="multi")

            uni_format['uni'] = st.radio("Select your uni", ["UON", "UNE"], horizontal=True)

            if uni_format['uni'] == "UON":
                uni_format['campus_code'] = st.text_input(
                    "Enter the campus you go to as shown on the spreadsheet. e.g. CAL/CC or Callaghan/Central Coast.")

            # verify_columns = st.button("Press to verify that your column settings are correct")

            st.info(f"Your settings right now:\n{uni_format}")

            if uni_format['day'] == 20:
                if os.path.exists("log.jccl"):
                    f = open("log.jccl", "r")
                    st.download_button("Download log file", data=f, file_name="log.jccl", use_container_width=True)
                else:
                    st.error("log not found")

            if uni_format['fields']:
                pass
            else:
                st.error("Fill your settings correctly")
                valid_custom = False
        else:
            uni_format = format_selection(uni_format_select)

    date_start, date_end = None, None

    valid_selection = False

    if file and valid_custom:
        # process file and determine which date ranges are valid
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        st.divider()
        st.subheader("Options")

        option = st.radio("Convert", ["All events", "Custom dates"], horizontal=True)

        campus = None
        pbl = None
        clin = None
        comm = None

        column1, column2 = st.columns(2)

        if option == "All events":
            date_start = datetime.date(1970, 1, 1)
            date_end = datetime.date(3000, 1, 1)
        elif option == "Custom dates":
            with column1:
                date_start = st.date_input("Date to import from (inclusive)", value="today", min_value=None,
                                           max_value=None,
                                           format="DD/MM/YYYY", key="dsi")
            with column2:
                date_end = st.date_input("Last date to import (inclusive)", value="today", min_value=None,
                                         max_value=None,
                                         format="DD/MM/YYYY", key="dei")


        if uni_format["uni"] == "UON":
            campus = st.selectbox("Campus", ["Callaghan", "Central Coast"])

        colpbl, colclin, colcomm = st.columns(3)

        if not st.toggle("Manual entry"):
            with colpbl:
                if "pbl" in uni_format["fields"]:
                    pbl = st.selectbox("PBL group", ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"), index=None, placeholder="Select PBL group")
                else:
                    pbl = "z"
            with colclin:
                if "clin" in uni_format["fields"]:
                    clin = st.selectbox("Clin group", ("1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16", "17", "18", "19", "20"), index=None, placeholder="Select clin group")
                else:
                    clin = "300"
            with colcomm:
                if "comm" in uni_format["fields"]:
                    comm = st.selectbox("Comm group", ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"), index=None, placeholder="Select comm group")
                else:
                    comm = "z"
        else:
            with colpbl:
                if "pbl" in uni_format["fields"]:
                    pbl = st.text_input("PBL group", placeholder="Select PBL group")
                else:
                    pbl = "z"
            with colclin:
                if "clin" in uni_format["fields"]:
                    clin = st.text_input("Clin group", placeholder="Select clin group")
                else:
                    clin = "300"
            with colcomm:
                if "comm" in uni_format["fields"]:
                    comm = st.text_input("Comm group", placeholder="Select comm group")
                else:
                    comm = "z"

        if option == "All events":
            st.text(f"Importing all available events from {file}")

        if clin and comm and pbl:
            valid_selection = True

        go = False


        if valid_selection and pbl:
            go = st.button("Start converting", use_container_width=True)
            st.text("Always check the output below for errors and double check calendar events with the spreadsheet on canvas. Scroll down to download the file when it's ready.")
        else:
            st.warning(f"Select your timetable.")

        if go and valid_selection:
            saved = process_xlsx(pbl.upper(), clin, comm, uni_format, campus, ws)
            converted = generate_cal(saved, date_start, date_end, uni_format)
            if os.path.exists("log.jccl"):
                with open("log.jccl", "w") as l:
                    log = f"{datetime.datetime.now()}: {uni_format['uni']} Calendar converted PBL: {pbl} clin: {clin} comm: {comm}\n\n"
                    l.write(log)

            if os.path.exists("calendar.ics"):
                f = open("calendar.ics", "r")

                st.download_button("Download ics file", data=f, file_name="Calendar.ics", use_container_width=True)

                st.text(
                    f"Import this file to your calendar app (google calendar works idk about the rest)\nAlways double check to see if events have been imported correctly. {converted} events were converted by the converter. If google calendar doesn't import the same number of events, check the spreadsheet. DM me @photogrudesh on Instagram if there are any issues.")



def process_xlsx(pbl, clin, comm, uni_format, campus, ws):
    events = []
    saved = []

    row_num = 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20, values_only=True):
        events.append(row)

    for i in events:
        i = list(i)
        keep = False
        group = str(i[uni_format["group"]]).strip()
        complex_grp = group.split(" ")[-1]

        if "ALL" in group:
            keep = True
        elif group.startswith("PBL"):
            if bool(re.search(rf'(?<!\w){re.escape(pbl)}(?!\w)', group)):
                keep = True
            if "-" in complex_grp:
                start_char = complex_grp.split("-")[0].strip()
                end_char = complex_grp.split("-")[1].strip()
                for x in [chr(c) for c in range(ord(start_char), ord(end_char) + 1)]:
                    if pbl == str(x):
                        keep = True
            elif "," in group:
                pbls = group.split(",")
                for f in pbls:
                    if f.strip() == pbl:
                        keep = True
        elif len(group) == 1:
            if group == pbl:
                keep = True

        elif group.startswith("Clin"):
            if bool(re.search(rf'(?<!\w){re.escape(clin)}(?!\w)', group)):
                keep = True
            if "-" in complex_grp:
                start, end = complex_grp.split("-")
                start_num = ""
                end_num = ""
                clin_groups = []
                for n in start:
                    if n.isdigit():
                        start_num += n
                start_num = int(start_num)
                for n in end:
                    if n.isdigit():
                        end_num += n
                end_num = int(end_num)
                for n in range(start_num, end_num + 1):
                    clin_groups.append(str(n))
                if clin in clin_groups:
                    keep = True

            elif "," in group:
                pbls = group.split(",")
                for f in pbls:
                    if f.strip() == pbl:
                        keep = True

        elif group.startswith("Comms"):
            if bool(re.search(rf'(?<!\w){re.escape(comm)}(?!\w)', group)):
                keep = True
            if "-" in complex_grp:
                start_char = complex_grp.split("-")[0].strip()
                end_char = complex_grp.split("-")[1].strip()
                for x in [chr(c) for c in range(ord(start_char), ord(end_char) + 1)]:
                    if pbl == str(x):
                        keep = True
            elif "," in group:
                comms = group.split(",")
                for f in comms:
                    if f.strip() == comm:
                        keep = True
        else:
            if "-" in complex_grp:
                start_char = complex_grp.split("-")[0].strip()
                end_char = complex_grp.split("-")[1].strip()
                for x in [chr(c) for c in range(ord(start_char), ord(end_char) + 1)]:
                    if pbl == str(x):
                        keep = True
            elif "," in group:
                pbls = group.split(",")
                for f in pbls:
                    if f.strip() == pbl:
                        keep = True


        if "zoom" in str(i[uni_format["venue"]]).lower():
            zoom_col = 8
            try:
                if uni_format["preset"] == "University of New England Y2 2026":
                    zoom_col = 7
                i[uni_format["venue"]] = ws.cell(row=row_num, column=zoom_col).hyperlink.target
                print(i[uni_format["venue"]])
            except AttributeError:
                print(f"No link for {i}")
                i[uni_format["venue"]] = str(i[uni_format["venue"]]) + ": Failed hyperlink extraction. Check spreadsheet."

        if uni_format["preset"] == "University of Newcastle Y2 2026":
            campus_code = None
            if campus == "Callaghan":
                campus_code = "CAL"
            elif campus == "Central Coast":
                campus_code = "CC"
            if campus_code not in str(i[uni_format["campus"]]):
                keep = False

        elif not uni_format["preset"]:
            campus_code = uni_format["campus_code"]
            if campus_code not in str(i[uni_format["campus"]]):
                keep = False

        if keep:
            saved.append(i)
        row_num += 1
    return saved


def generate_cal(events, date_start, date_end, uni_format):
    cal = Calendar()
    cal['version'] = '2.0'
    cal.add("prodid", "-//photogrudesh//JMPCalendarConverter//EN")
    cal.add("summary", "JMP schedule")
    converted = 0

    with st.status(f"Importing events...", expanded=True):
        for i in events:
            try:
                if date_start - datetime.timedelta(days=1) < i[uni_format["date"]].date() < date_end + datetime.timedelta(days=1):
                    event = Event()
                    no_time = False

                    try:
                        start, end = convert_datetime(i[uni_format["date"]], i[uni_format["time"]])
                        event.add('dtstart', start)
                        event.add('dtend', end)
                    except TypeError:
                        no_time = True

                    if i[uni_format["attendance"]] is None:
                        attendance = "N/A"
                    else:
                        attendance = i[uni_format["attendance"]]

                    desc = f"{i[uni_format["type"]]}: {i[uni_format["domain"]]}\n{i[uni_format["venue"]]}\nStudents: {i[uni_format["group"]]}\nAttendance: {attendance}\nStaff: {i[uni_format["staff"]]}\nUpdates: {i[uni_format["updates"]]}"

                    if no_time:
                        event.add('dtstart', i[uni_format["date"]])
                        event.add('dtend', i[uni_format["date"]] + datetime.timedelta(days=1))

                    event.add('summary', i[uni_format["session"]])
                    event.add("description", desc)
                    cal.add_component(event)
                    if not no_time:
                        with st.status(f"{start}: " + i[uni_format["session"]].replace('\n', ' ')):
                            st.write(desc)
                    else:
                        with st.status(i[uni_format["session"]].replace('\n', ' ')):
                            st.write(desc)
                    converted += 1
            except AttributeError as e:
                print(e)
                st.error(f"Something went wrong. Check the spreadsheet for information on {i}\n{e}")

    with open("calendar.ics", "wb") as f:
        f.write(cal.to_ical())
    return converted


def convert_datetime(date, time):
    time = time.replace(".", ":").replace("md", "pm")

    times = time.split(" - ")

    try:
        cal = pdt.Calendar()
        time_struct, parse_status = cal.parse(times[0])
        start_time = datetime.datetime(*time_struct[:6]).time()
        cal = pdt.Calendar()
        time_struct, parse_status = cal.parse(times[1])
        end_time = datetime.datetime(*time_struct[:6]).time()

        start = datetime.datetime.combine(date, start_time)
        end = datetime.datetime.combine(date, end_time)

    except IndexError:
        return date
    return start, end


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main_ui()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
